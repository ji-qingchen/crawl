import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border, Font

#定义填充色，红色：困难，黄色：中等，绿色：简单
#头：黑色，line1: 蓝，line2: 紫
#字体：白色，黑色
head_fill = PatternFill('solid', fgColor = '636363') 
line1_fill = PatternFill('solid', fgColor = '6600cc')
line2_fill = PatternFill('solid', fgColor = '00ccff')

dif_fill = PatternFill('solid', fgColor = 'ff6600')
fit_fill = PatternFill('solid', fgColor = 'fff000')
eas_fill = PatternFill('solid', fgColor = '66cc00')

#对齐方式
align = Alignment( horizontal = 'center', vertical = 'center')

#边框样式
side = Side('thin')

#边框类型
border = Border(top = side, bottom = side, left = side, right = side)

#字体
color1 = 'ffffff' #白
color2 = '000000'
a = '宋体' 
font_head = Font(name='宋体',size=14, bold=True, italic=False, color=color1)   #bold指是否加粗,italic指是否倾斜
font_sec = Font(name='宋体',size=12, bold=True, italic=False, color=color1)
font_other = Font(name='宋体',size=10, bold=False, italic=False, color=color2)

#路径
path = r'C:\Users\姬清晨\Desktop\正在做\学习\算法\春招班五毒神掌完整版 - 副本.xlsx'

#打开工作簿
wb = load_workbook(path)
#打开工作表
ws = wb.get_sheet_by_name('week2')

#写入内容
rows = [['题号', 'LeetCode 链接', '知识点', '一', '二', '三', '四', '五','注'], ['第二周','','','','','','','','注：绿色为简单\n黄色为中等\n红色为困难\n'], ['实战']]
for row in rows:
    ws.append(row)
ws['A19'] = '实战'

#设置列宽：
ws.column_dimensions['A'].width = 10.21
ws.column_dimensions['B'].width = 19.38
ws.column_dimensions['C'].width = 44.54
ws.column_dimensions['D'].width = 10.54
ws.column_dimensions['E'].width = 4.04
ws.column_dimensions['F'].width = 4.04
ws.column_dimensions['G'].width = 4.04
ws.column_dimensions['H'].width = 4.04
ws.column_dimensions['I'].width = 21


#设置行宽
ws.row_dimensions[1].height = 18
ws.row_dimensions[2].height = 70
ws.row_dimensions[3].height = 48



for i in range(4, 18):
    ws.row_dimensions[i].height = 26

ws.row_dimensions[19].height = 48

for j in range(20, 35):
    ws.row_dimensions[j].height = 26

#设置head行
for cell in ws[1]:
    cell.fill = head_fill
    
#第二、三行
ws['A2'].fill = line1_fill
ws['I2'].fill = fit_fill
ws['A3'].fill = line2_fill
ws['A19'].fill = line2_fill

#设置边框，对齐方式
#rowe = ws.max_row
for row2 in ws.iter_rows(max_col = 9, max_row=17):
    for cell2 in row2:
        
        cell2.alignment = align
        cell2.border = border
        cell2.font = font_other

for row3 in ws.iter_rows(max_col=9,min_row = 19, max_row=34):
    for cell3 in row3:
     
        cell3.alignment = align
        cell3.border = border
        cell3.font = font_other
#设置字体
for cell31 in ws[1]:
    cell31.font = font_head

for cell4 in ws[2]:
    cell4.font = font_sec

for cell4 in ws[3]:
    cell4.font = font_sec

for cell4 in ws[19]:
    cell4.font = font_sec


wb.save(path)






